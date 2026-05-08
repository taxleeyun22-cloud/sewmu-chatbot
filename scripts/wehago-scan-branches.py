#!/usr/bin/env python3
"""위하고 export — 옆커폰 / 옆커호텔 본점 row 찾기 + 법인등록번호 매칭 가능성 확인.

사장님 명령 (2026-05-08): "1~5번은 주식회사 옆커폰 지점, 6번은 옆커호텔 지점.
법인등록번호가 알아서 잘 넣을 수 있겠어?"
"""
import sys, os
import openpyxl


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(5, ws.max_row + 1, 2):
        row1 = [ws.cell(r, c).value for c in range(1, 9)]
        row2 = [ws.cell(r+1, c).value for c in range(1, 9)] if r+1 <= ws.max_row else [None]*8
        if not row1[0]:
            continue
        rows.append({
            'no': str(row1[0]) if row1[0] is not None else '',
            'company': str(row1[2]).strip() if row1[2] else '',
            'corp_or_indiv': str(row1[4]).strip() if row1[4] else '',
            'biz_no': str(row1[5]).strip() if row1[5] else '',
            'resident_or_corp_no': str(row1[6]).strip() if row1[6] else '',
            'phone': str(row1[7]).strip() if row1[7] else '',
            'ceo': str(row2[2]).strip() if row2[2] else '',
        })
    return rows


def main():
    path = r'C:\Users\user\Desktop\taxtrade_최종정리.xlsx'
    rows = parse_xlsx(path)
    print(f'>> 총 {len(rows)} row\n')

    # 1. 회사명에 "옆커폰" 들어간 row 전부
    print('=' * 70)
    print('🔍 회사명에 "옆커폰" 들어간 row:')
    print('=' * 70)
    yeokpon = [r for r in rows if '옆커폰' in r['company']]
    for r in yeokpon:
        marker = '⭐' if r['corp_or_indiv'] == '법인' and r['resident_or_corp_no'] else \
                 '⚠️ ' if not r['resident_or_corp_no'] else '  '
        print(f'  {marker} [{r["corp_or_indiv"]}] {r["company"]}')
        print(f'        대표: {r["ceo"]} / 사업자번호: {r["biz_no"]}')
        print(f'        법인/주민번호: "{r["resident_or_corp_no"]}"')
        print()

    # 2. "옆커호텔" 키워드 row
    print('=' * 70)
    print('🔍 회사명에 "옆커호텔" 들어간 row:')
    print('=' * 70)
    yeokhotel = [r for r in rows if '옆커호텔' in r['company']]
    for r in yeokhotel:
        marker = '⭐' if r['corp_or_indiv'] == '법인' and r['resident_or_corp_no'] else \
                 '⚠️ ' if not r['resident_or_corp_no'] else '  '
        print(f'  {marker} [{r["corp_or_indiv"]}] {r["company"]}')
        print(f'        대표: {r["ceo"]} / 사업자번호: {r["biz_no"]}')
        print(f'        법인/주민번호: "{r["resident_or_corp_no"]}"')
        print()

    # 3. 회사명에 "[...지점]" 또는 "[옆커..." 표기 있는 row 전부
    print('=' * 70)
    print('🔍 지점 표기 패턴 분석 — "[" 가 회사명에 있는 모든 법인 row:')
    print('=' * 70)
    bracket_corps = [r for r in rows if '[' in r['company'] and r['corp_or_indiv'] == '법인']
    for r in bracket_corps:
        marker = '✅' if r['resident_or_corp_no'] else '⚠️ '
        print(f'  {marker} {r["company"]}')
        print(f'        대표: {r["ceo"]} / 사업자번호: {r["biz_no"]} / 법인번호: "{r["resident_or_corp_no"]}"')
    print()

    # 4. 법인 row 중 corp_no 비어있는거만 다시
    print('=' * 70)
    print('⚠️  법인 row 중 법인등록번호 비어있음:')
    print('=' * 70)
    empty_corp = [r for r in rows if r['corp_or_indiv'] == '법인' and not r['resident_or_corp_no']]
    for r in empty_corp:
        # 회사명에서 본점 키워드 추출 시도
        guess = ''
        if '[옆커폰' in r['company']:
            guess = '옆커폰주식회사 (또는 비슷)'
        elif '[옆커호텔' in r['company']:
            guess = '옆커호텔 (또는 비슷)'
        print(f'  - {r["company"]}')
        print(f'      추측 본점: {guess}')

    # 5. 자동 매칭 시뮬레이션
    print()
    print('=' * 70)
    print('💡 자동 매칭 가능성:')
    print('=' * 70)

    # 옆커폰 법인 + 법인등록번호 있는 row 찾기
    yeokpon_corp_rows = [r for r in yeokpon if r['corp_or_indiv'] == '법인' and r['resident_or_corp_no']]
    print(f'\n  옆커폰 법인 + 법인번호 있는 row: {len(yeokpon_corp_rows)}건')
    if yeokpon_corp_rows:
        # 가장 단순한 회사명 (= 본점 후보)
        sorted_by_len = sorted(yeokpon_corp_rows, key=lambda x: len(x['company']))
        print('  가장 단순한 회사명 (본점 후보):')
        for r in sorted_by_len[:3]:
            print(f'    · {r["company"]} (대표 {r["ceo"]}, 법인번호 {r["resident_or_corp_no"]})')

    yeokhotel_corp_rows = [r for r in yeokhotel if r['corp_or_indiv'] == '법인' and r['resident_or_corp_no']]
    print(f'\n  옆커호텔 법인 + 법인번호 있는 row: {len(yeokhotel_corp_rows)}건')
    if yeokhotel_corp_rows:
        sorted_by_len = sorted(yeokhotel_corp_rows, key=lambda x: len(x['company']))
        print('  가장 단순한 회사명 (본점 후보):')
        for r in sorted_by_len[:3]:
            print(f'    · {r["company"]} (대표 {r["ceo"]}, 법인번호 {r["resident_or_corp_no"]})')

    return 0


if __name__ == '__main__':
    sys.exit(main())
