#!/usr/bin/env python3
"""
위하고 export 엑셀 → backend preview/commit 호출 스크립트
사장님 명령 (2026-05-08): "엑셀 줄게 → 자동 등록 + 롤백 가능"

사용법:
  python scripts/wehago-import.py --file taxtrade_최종정리.xlsx --key ADMIN_KEY [--commit]

옵션:
  --file        위하고 export 엑셀 경로
  --key         ADMIN_KEY
  --base-url    backend URL (default: https://sewmu-chatbot.pages.dev)
  --commit      preview 만 보지 않고 자동 commit (default: preview only)
  --batch-id    이미 preview 된 batch 의 commit (commit 단독 호출)

흐름:
  1. 엑셀 → JSON rows 파싱 (1 거래처 = 2 row 형식)
  2. POST /api/admin-bulk-import-clients?action=preview → batch_id, 결과 보고
  3. (사장님 OK 시) POST ?action=commit&batch_id=N → 실제 INSERT
  4. 결과 보고 (롤백은 admin UI 에서)
"""
import sys, os, argparse, json, re, urllib.request, urllib.error
import openpyxl


def extract_branch_tag(name):
    """회사명에서 [...] 안 태그 추출 (지점 표시).
    예: '두산동지점 [옆커폰/유킹지점]' -> '옆커폰/유킹지점'"""
    if not name:
        return None
    m = re.search(r'\[([^\]]+)\]', name)
    return m.group(1).strip() if m else None


def enrich_empty_corp_no(rows):
    """법인 row 중 법인등록번호 비어있으면 같은 [태그] 의 다른 row 법인번호로 자동 채움.
    사장님 명령 (2026-05-08): "1~5번은 옆커폰주식회사 지점, 6번은 옆커호텔 지점.
    법인등록번호가 알아서 잘 넣을수 있겠어?"

    알고리즘:
      1) 회사명에서 [...] 태그 추출
      2) 같은 태그 가진 다른 row 의 법인번호 그룹화
      3) 단일 법인번호로 통일되면 빈 row 도 그 번호로 채움
      4) 다중 / 매칭 0 → enrichment 안 함, warning 만 표시

    추가 변환:
      - '옆커폰지점/유킹' (variant) → '옆커폰/유킹지점' 정규화 (같은 그룹 인식)
    """
    def normalize_tag(t):
        if not t:
            return t
        # 동일 그룹 variant 정규화
        if t in ('옆커폰지점/유킹', '옆커폰/유킹', '옆커폰/유킹지점'):
            return '옆커폰/유킹지점'
        return t

    # 1. 태그 → 법인번호 set 매핑 (법인 + 법인번호 있는 row 만)
    tag_to_corp_nos = {}
    for r in rows:
        if r['corp_or_indiv'] != '법인' or not r['resident_or_corp_no']:
            continue
        tag = normalize_tag(extract_branch_tag(r['company']))
        if not tag:
            continue
        tag_to_corp_nos.setdefault(tag, set()).add(r['resident_or_corp_no'])

    # 2. 빈 법인 row 채우기
    enriched = []
    skipped = []
    for r in rows:
        if r['corp_or_indiv'] != '법인' or r['resident_or_corp_no']:
            continue
        tag = normalize_tag(extract_branch_tag(r['company']))
        if not tag:
            skipped.append({'company': r['company'], 'reason': '태그 추출 실패'})
            continue
        candidates = tag_to_corp_nos.get(tag, set())
        if len(candidates) == 1:
            corp_no = next(iter(candidates))
            r['resident_or_corp_no'] = corp_no
            enriched.append({'company': r['company'], 'tag': tag, 'corp_no': corp_no, 'via': 'tag'})
            continue

        # Fallback: 같은 [태그] row 0건 / 다중 → 태그에서 키워드 추출 후 회사명 검색
        # 예: '옆커호텔지점' → 키워드 '옆커호텔' → 회사명에 '옆커호텔' 들어간 법인 row 의 법인번호
        keyword = re.sub(r'(지점|지사|영업소|센터점|/[^/]*$|/[^/]*지점)', '', tag).strip()
        if keyword and len(keyword) >= 2:
            kw_candidates = set()
            for other in rows:
                if other is r:
                    continue
                if other['corp_or_indiv'] != '법인' or not other['resident_or_corp_no']:
                    continue
                if keyword in other['company']:
                    kw_candidates.add(other['resident_or_corp_no'])
            if len(kw_candidates) == 1:
                corp_no = next(iter(kw_candidates))
                r['resident_or_corp_no'] = corp_no
                enriched.append({'company': r['company'], 'tag': tag, 'corp_no': corp_no, 'via': 'keyword=' + keyword})
                continue
            elif len(kw_candidates) > 1:
                skipped.append({'company': r['company'], 'tag': tag, 'reason': f'키워드 "{keyword}" → 후보 {len(kw_candidates)}개 (수동 결정 필요)'})
                continue

        if len(candidates) == 0:
            skipped.append({'company': r['company'], 'tag': tag, 'reason': '같은 태그 row 0건 + 키워드 매칭도 실패'})
        else:
            skipped.append({'company': r['company'], 'tag': tag, 'reason': f'같은 태그 다른 법인번호 {len(candidates)}개'})

    return enriched, skipped


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    # 1 거래처 = 2 row, 시작 row 5 (header 4줄)
    for r in range(5, ws.max_row + 1, 2):
        row1 = [ws.cell(r, c).value for c in range(1, 9)]
        row2 = [ws.cell(r+1, c).value for c in range(1, 9)] if r+1 <= ws.max_row else [None]*8
        if not row1[0]:  # 번호 없음 (빈 row)
            continue
        rows.append({
            'no': str(row1[0]) if row1[0] is not None else '',
            'code': str(row1[1]) if row1[1] is not None else '',
            'company': str(row1[2]).strip() if row1[2] else '',
            'type1': str(row1[3]).strip() if row1[3] else '',  # 업태
            'corp_or_indiv': str(row1[4]).strip() if row1[4] else '',  # 법인/개인
            'biz_no': str(row1[5]).strip() if row1[5] else '',
            'resident_or_corp_no': str(row1[6]).strip() if row1[6] else '',
            'phone': str(row1[7]).strip() if row1[7] else '',
            'ceo': str(row2[2]).strip() if row2[2] else '',
            'industry': str(row2[3]).strip() if row2[3] else '',  # 종목
            'address': str(row2[4]).strip() if row2[4] else '',
            'home_addr': str(row2[6]).strip() if row2[6] else '',
            'tax_office': '',  # 위하고 export 에 별도 컬럼 없음, 주소 분석 필요
        })

    # 사장님 명령 (2026-05-08): 법인 row 빈 법인번호 자동 채움
    enriched, skipped = enrich_empty_corp_no(rows)
    if enriched or skipped:
        print(f'>> 법인등록번호 자동 enrichment: {len(enriched)}건 채움 / {len(skipped)}건 skip')
        for e in enriched:
            print(f'   ✅ {e["company"]} → {e["corp_no"]} (태그: {e["tag"]})')
        for s in skipped:
            print(f'   ⚠️  {s["company"]} → skip ({s["reason"]})')

    return rows


def call_api(base_url, key, action, body):
    url = f"{base_url}/api/admin-bulk-import-clients?action={action}&key={key}"
    req = urllib.request.Request(
        url,
        data=json.dumps(body, ensure_ascii=False).encode('utf-8'),
        headers={
            'Content-Type': 'application/json; charset=utf-8',
            'User-Agent': 'Mozilla/5.0 (sewmu-import-script) AppleWebKit/537.36',
            'Accept': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            return json.loads(r.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8') if e.fp else ''
        return {'ok': False, 'error': f'HTTP {e.code}: {body[:500]}'}
    except Exception as e:
        return {'ok': False, 'error': f'fetch fail: {e}'}


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--file', required=False, help='위하고 export xlsx')
    p.add_argument('--key', required=True, help='ADMIN_KEY')
    p.add_argument('--base-url', default='https://sewmu-chatbot.pages.dev')
    p.add_argument('--commit', action='store_true', help='preview 후 자동 commit')
    p.add_argument('--batch-id', type=int, help='기존 preview batch_id (commit 단독 호출)')
    args = p.parse_args()

    if args.batch_id and not args.file:
        # commit only mode
        print(f'>> commit batch_id={args.batch_id}')
        result = call_api(args.base_url, args.key, 'commit', {'batch_id': args.batch_id})
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0 if result.get('ok') else 1

    if not args.file:
        print('--file 필요 (또는 --batch-id 로 commit only)')
        return 1
    if not os.path.exists(args.file):
        print(f'파일 없음: {args.file}')
        return 1

    print(f'>> parsing {args.file}')
    rows = parse_xlsx(args.file)
    print(f'>> {len(rows)} 거래처 파싱 완료')
    print(f'   - 개인: {sum(1 for r in rows if r["corp_or_indiv"] == "개인")}')
    print(f'   - 법인: {sum(1 for r in rows if r["corp_or_indiv"] == "법인")}')

    print(f'>> POST preview → {args.base_url}')
    preview_result = call_api(args.base_url, args.key, 'preview', {
        'source_file': os.path.basename(args.file),
        'rows': rows,
    })
    if not preview_result.get('ok'):
        print(f'>> preview 실패: {preview_result.get("error")}')
        return 1

    pv = preview_result['preview']
    print()
    print('=' * 60)
    print(f'📊 PREVIEW (batch_uuid: {pv["batch_uuid"]}, batch_id: {pv["batch_id"]})')
    print('=' * 60)
    print(f'총 거래처: {pv["total_rows"]}')
    print(f'사용자:')
    print(f'  - 신규 INSERT 예정: {pv["users"]["new"]}')
    print(f'  - 기존 매칭 (dedup): {pv["users"]["matched"]}')
    print(f'  - 빈 컬럼 enrichment: {pv["users"]["enriched"]}')
    print(f'사업장:')
    print(f'  - 신규 INSERT 예정: {pv["businesses"]["new"]}')
    print(f'  - dedup (이미 있음): {pv["businesses"]["dedup"]}')
    print(f'본점·지점 그룹: {pv["branch_groups"]["auto_ok"]} 자동 OK / {pv["branch_groups"]["needs_decision"]} 결정 필요')
    if pv.get('warnings'):
        print(f'⚠️  경고 ({len(pv["warnings"])}):')
        for w in pv['warnings'][:10]:
            print(f'  - {w}')
    print()
    print('🔍 본점·지점 그룹 상세:')
    for grp in (pv.get('branch_group_list') or [])[:10]:
        marker = '✅' if grp['status'] == 'auto_ok' else '⚠️'
        main_label = grp['main_row']['company'] if grp.get('main_row') else '⚠️ 본점 row 없음'
        print(f'  {marker} {grp["corp_no"]} ({grp["group_size"]}개) — 본점: {main_label}')
        for r in grp['all_rows'][:5]:
            print(f'     · {r["company"]} ({r["biz_no"]})')

    if not args.commit:
        print()
        print(f'== preview 완료. 실제 INSERT 하려면 --commit 옵션 또는 별도 호출:')
        print(f'   python {sys.argv[0]} --batch-id {pv["batch_id"]} --key {args.key}')
        return 0

    print()
    print(f'>> POST commit batch_id={pv["batch_id"]}')
    commit_result = call_api(args.base_url, args.key, 'commit', {'batch_id': pv['batch_id']})
    if not commit_result.get('ok'):
        print(f'>> commit 실패: {commit_result.get("error")}')
        return 1
    print()
    print('=' * 60)
    print(f'✅ COMMIT 완료')
    print('=' * 60)
    print(json.dumps(commit_result, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    sys.exit(main())
