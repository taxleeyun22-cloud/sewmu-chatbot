#!/usr/bin/env python3
"""위하고 export — 주민번호 비어있는 row 전수조사.

사장님 명령 (2026-05-08): "주민번호 비어있는 row 누군지 전수조사 먼저"
이름 + 주민번호 dedup 룰 도입 전 확인 — 주민번호 없는 사람들은 어떻게 처리할지.
"""
import sys, os
import openpyxl


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(5, ws.max_row + 1, 2):
        row1 = [ws.cell(r, c).value for c in range(1, 9)]
        row2 = [ws.cell(r+1, c).value for c in range(1, 9)] if r+1 <= ws.max_row else [None]*8
        if not row1[0]:
            continue
        rows.append({
            'no': str(row1[0]) if row1[0] is not None else '',
            'code': str(row1[1]) if row1[1] is not None else '',
            'company': str(row1[2]).strip() if row1[2] else '',
            'type1': str(row1[3]).strip() if row1[3] else '',
            'corp_or_indiv': str(row1[4]).strip() if row1[4] else '',
            'biz_no': str(row1[5]).strip() if row1[5] else '',
            'resident_or_corp_no': str(row1[6]).strip() if row1[6] else '',
            'phone': str(row1[7]).strip() if row1[7] else '',
            'ceo': str(row2[2]).strip() if row2[2] else '',
            'industry': str(row2[3]).strip() if row2[3] else '',
            'address': str(row2[4]).strip() if row2[4] else '',
            'home_addr': str(row2[6]).strip() if row2[6] else '',
        })
    return rows


def is_personal_rrn(rrn):
    """주민번호인지 (법인등록번호 아님) — 13자리 - 7자리 패턴 또는 13자리 numeric"""
    if not rrn:
        return False
    # 하이픈 제거
    clean = rrn.replace('-', '').replace(' ', '')
    if len(clean) != 13 or not clean.isdigit():
        return False
    # 7번째 자리 (back 첫자리) 가 1,2,3,4,5,6,7,8 → 주민번호 가능
    back_first = clean[6]
    return back_first in '12345678'


def main():
    path = r'C:\Users\user\Desktop\taxtrade_최종정리.xlsx'
    if not os.path.exists(path):
        print(f'파일 없음: {path}')
        return 1

    rows = parse_xlsx(path)
    print(f'>> 총 {len(rows)} 거래처 파싱됨')
    print()

    # 분류
    empty_rrn = []           # resident_or_corp_no 자체 비어있음
    corp_no = []             # 법인등록번호 (corp_or_indiv == '법인')
    has_rrn = []             # 주민번호 있음
    weird = []               # 어디에도 안 맞는 케이스

    for r in rows:
        rrn = r['resident_or_corp_no']
        kind = r['corp_or_indiv']
        if not rrn:
            empty_rrn.append(r)
        elif kind == '법인':
            corp_no.append(r)
        elif is_personal_rrn(rrn):
            has_rrn.append(r)
        else:
            weird.append(r)

    print('=' * 70)
    print(f'분류 결과:')
    print(f'  - 주민번호 있음 (개인): {len(has_rrn)}건')
    print(f'  - 법인등록번호 있음 (법인): {len(corp_no)}건')
    print(f'  - 주민번호 비어있음:       {len(empty_rrn)}건  ⭐')
    print(f'  - 어디에도 안 맞음:        {len(weird)}건')
    print('=' * 70)
    print()

    if empty_rrn:
        print('🚨 주민번호 비어있는 row 전수 (사장님 검토 대상):')
        print('-' * 70)
        for i, r in enumerate(empty_rrn, 1):
            print(f'{i:3}. [{r["corp_or_indiv"] or "?"}] {r["company"]}')
            print(f'      대표: {r["ceo"]} / 사업자번호: {r["biz_no"]} / 전화: {r["phone"]}')
            print(f'      업태: {r["type1"]} / 종목: {r["industry"]}')
            print(f'      주소: {r["address"]}')
            print()

    if weird:
        print()
        print('⚠️ 어디에도 안 맞는 row (검토 필요):')
        print('-' * 70)
        for i, r in enumerate(weird, 1):
            print(f'{i:3}. [{r["corp_or_indiv"] or "?"}] {r["company"]}')
            print(f'      대표: {r["ceo"]} / 주민번호 자리: "{r["resident_or_corp_no"]}"')
            print(f'      사업자번호: {r["biz_no"]} / 전화: {r["phone"]}')
            print()

    # 같은 이름인데 주민번호 비어있는 사람들 — dedup 영향 큼
    print()
    print('=' * 70)
    print('🔍 dedup 영향 분석 (이름+주민번호 룰 기준):')
    print('=' * 70)

    # has_rrn 안 같은 이름 + 같은 주민번호
    name_rrn_map = {}
    for r in has_rrn:
        key = (r['ceo'], r['resident_or_corp_no'])
        name_rrn_map.setdefault(key, []).append(r)

    same_person_multi_biz = {k: v for k, v in name_rrn_map.items() if len(v) > 1}
    print(f'\n  이름+주민번호 같음 → 같은 사람 (1 user, N businesses):')
    print(f'  → {len(same_person_multi_biz)} 명 / 총 {sum(len(v) for v in same_person_multi_biz.values())} row')
    if same_person_multi_biz:
        for (name, rrn), rs in list(same_person_multi_biz.items())[:20]:
            companies = ', '.join([x['company'] for x in rs])
            print(f'    · {name} ({rrn[:6]}-*******): {len(rs)}개 업체 — {companies}')

    # has_rrn 안 같은 이름인데 주민번호 다름 (동명이인 가능)
    name_only_map = {}
    for r in has_rrn:
        name_only_map.setdefault(r['ceo'], set()).add(r['resident_or_corp_no'])
    homonyms = {k: v for k, v in name_only_map.items() if len(v) > 1}
    print(f'\n  같은 이름인데 주민번호 다름 (동명이인 가능):')
    print(f'  → {len(homonyms)} 명')
    if homonyms:
        for name, rrns in list(homonyms.items())[:20]:
            print(f'    · {name}: 주민번호 {len(rrns)}개')
            for rrn in rrns:
                rs = [x for x in has_rrn if x['ceo'] == name and x['resident_or_corp_no'] == rrn]
                companies = ', '.join([x['company'] for x in rs])
                print(f'        - {rrn[:6]}-******* → {len(rs)}개 업체: {companies}')

    # empty_rrn 중 같은 이름인 사람 (주민번호 없으니 dedup 불가)
    empty_name_map = {}
    for r in empty_rrn:
        empty_name_map.setdefault(r['ceo'], []).append(r)
    empty_dups = {k: v for k, v in empty_name_map.items() if len(v) > 1 and k}
    print(f'\n  주민번호 비어있는데 같은 이름:')
    print(f'  → {len(empty_dups)} 명')
    if empty_dups:
        for name, rs in empty_dups.items():
            companies = ', '.join([x['company'] for x in rs])
            print(f'    · {name}: {len(rs)}개 업체 — {companies}')

    print()
    print('=' * 70)
    print('💡 결론:')
    print('=' * 70)
    print(f'  - 308 row 중 user 예상: {len(same_person_multi_biz)}명(중복묶임) + 단일')
    print(f'  - 주민번호 비어있는 사람 처리 정책 결정 필요:')
    print(f'    (a) 이름만으로 매칭 — 동명이인 위험')
    print(f'    (b) 주민번호 없으면 무조건 신규 — 매번 새 user (중복 user 양산 가능)')
    print(f'    (c) skip — import 안 함, 사장님이 admin 에서 수동')
    return 0


if __name__ == '__main__':
    sys.exit(main())
